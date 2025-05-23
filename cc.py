# import openpyxl
# from datetime import datetime

# # Define file path (make sure the file exists or create it)
# file_path = "data.xlsx"

# # Function to append data to the Excel file
# def append_to_excel(mobile_numbers, class_name):
#     # Get current time
#     time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

#     # Check if the file exists
#     try:
#         # Open the existing workbook
#         wb = openpyxl.load_workbook(file_path)
#     except FileNotFoundError:
#         # If file does not exist, create a new workbook and add header
#         wb = openpyxl.Workbook()
#         sheet = wb.active
#         sheet.append(['Mobile Number', 'Class Name', 'Time Now'])

#     # Get the active sheet
#     sheet = wb.active

#     # Loop through the list of mobile numbers and append each to the Excel sheet
#     for mobile_number in mobile_numbers:
#         sheet.append([mobile_number, class_name, time_now])

#     # Save the workbook
#     wb.save(file_path)
#     print(f"Data appended: {len(mobile_numbers)} entries for class {class_name}")

import openpyxl
from datetime import datetime

# Define file path (make sure the file exists or create it)
file_path = "data.xlsx"

# Function to append data to the Excel file
def append_to_excel(mobile_numbers, class_name, website):
    # Get current time
    time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print("a")
    # Check if the file exists
    try:
        # Open the existing workbook
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        # Check if the header exists, otherwise create it
        if sheet.max_row == 1 and sheet.cell(row=1, column=1).value != "Mobile Number":
            sheet.append(['Mobile Number', 'Class Name', 'Website', 'Time Now'])
    except FileNotFoundError:
        # If file does not exist, create a new workbook and add header
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Mobile Number', 'Class Name', 'Website', 'Time Now'])
        print("b")

    # Loop through the list of mobile numbers and append each to the Excel sheet
    for mobile_number in mobile_numbers:
        sheet.append([mobile_number, class_name, website, time_now])

    # Save the workbook
    wb.save(file_path)
    print(f"Data appended: {len(mobile_numbers)} entries for class {class_name} with website {website}")


